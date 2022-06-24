

from googletrans import Translator
from openpyxl import Workbook,load_workbook
wb = load_workbook("data.xlsx")
ws = wb.active

wb2 = Workbook()
ws2 = wb2.active
def cevir(cumle,konu):
    
    translator = Translator()
    tr=cumle
    print(konu)
    ws2.append([cumle,konu]) 
    tr2=""
    i=1
    while(i<5):
        
        txt = translator.translate(tr, src='tr', dest='en').text
        tr = translator.translate(txt, src='en', dest='tr').text        
        if(tr==tr2):
            break
        elif(tr==cumle):
            break
        else:
            ws2.append([tr,konu]) 
        print(str(i)+":"+txt)
        print(str(i)+":"+tr)
        tr2=tr
        i=i+1
i=1
for x in ws:
    
    cevir(ws[("A"+str(i))].value,ws[("B"+str(i))].value)
    print("____________________")
    i=i+1
wb2.save("data2.xlsx")
        
